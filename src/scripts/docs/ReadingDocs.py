import openpyxl
import pdfplumber

from src.scripts.formater import Formater

class ReadingDocs:
    def __init__(self):
        pass

    def leerXLS(self, ruta: str):
        """
        Reads an Excel file and extracts the data from the first sheet.

        This function opens an Excel file, extracts the data from the first sheet,
        and formats the data into a list of dictionaries using the Formater class.

        Args:
            ruta (str): The file path to the Excel document.

        Returns:
            list[dict]: A list of dictionaries, each representing a row from the 
            Excel table with column names as keys.
        """
        lector = openpyxl.load_workbook(ruta).active
        nombreColumnas = [lector.cell(row=1, column=i).value for i in range(1, lector.max_column + 1)]
        return Formater().formatoJSONDesdeXLS(nombreColumnas, lector)

    def leerPDF(self, ruta: str):
        """
        Reads a PDF file and extracts the first page's table data.

        This function opens a PDF file, extracts the table from the first page,
        and formats the data into a list of dictionaries using the Formater class.

        Args:
            ruta (str): The file path to the PDF document.

        Returns:
            list[dict]: A list of dictionaries, each representing a row from the 
            PDF table with column names as keys.
        """

        with pdfplumber.open(ruta) as pdf:
            preTabla = pdf.pages[0].extract_table()
            tabla = []
            for fila in preTabla:
                if None in fila:
                    fila.insert(fila.index(None), '')
                    fila.remove(None)
                tabla.append(fila)
            nombreColumnas = tabla[0]
            preValores = tabla[1:]
        return Formater().formatoJSONDesdePDF(nombreColumnas, preValores)