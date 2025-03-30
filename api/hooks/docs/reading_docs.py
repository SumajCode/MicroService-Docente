import openpyxl
import pdfplumber

from hooks.db.formater import format_sql_insert

def read_xls(path: str):
  """
  Reads an Excel file and creates a SQL insert statement from the data it contains

  :param path: Path to the Excel file
  :return: None
  """
  reader = openpyxl.load_workbook(path).active
  columns_name = [reader.cell(row=1, column=i).value for i in range(1, reader.max_column + 1)]
  values = []
  values_into = []
  for j in range(2, reader.max_row + 1):
    for i in range(1, reader.max_column + 1):
      value = reader.cell(row=j, column=i).value
      values_into.append("\""+str(value)+"\"" if isinstance(value,str) else str(int(value)))
    values.append(values_into)
    values_into = []
  return format_sql_insert('tabla', columns_name, values)

def read_pdf(path: str):
  """
  Reads a PDF file and creates a SQL insert statement from the data it contains

  :param path: Path to the PDF file
  :return: None
  """
  
  with pdfplumber.open(path) as pdf:
    pretable = pdf.pages[0].extract_table()
    table = []
    for row in pretable:
      if None in row:
        row.insert(row.index(None), '')
        row.remove(None)
      table.append(row)
    columns = table[0]
    prevalues = [row for row in table[1:]]
    values = []
    for row in prevalues:
      prerow = []
      for value in row:
        prerow.append(str(int(value)) if value.isdigit() else "\""+str(value)+"\"")
      values.append(prerow)
  return format_sql_insert(table="tabla", columns=columns, values=values)