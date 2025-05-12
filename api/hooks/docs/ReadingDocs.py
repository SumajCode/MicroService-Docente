import openpyxl
import pdfplumber

from orm.formater import format_sql_insert

def leerXLS(ruta: str):
  """
  Reads an Excel file and creates a SQL insert statement from the data it contains

  :param path: path to the Excel file
  :return: None
  """
  lector = openpyxl.load_workbook(ruta).active
  nombreColumnas = [lector.cell(row=1, column=i).value for i in range(1, lector.max_column + 1)]
  valoresColumna = []
  valoresIntro = []
  for j in range(2, lector.max_row + 1):
    for i in range(1, lector.max_column + 1):
      value = lector.cell(row=j, column=i).value
      valoresIntro.append("\""+str(value)+"\"" if isinstance(value,str) else str(int(value)))
    valoresColumna.append(valoresIntro)
    valoresIntro = []
  return format_sql_insert('tabla', nombreColumnas, valoresColumna)

def leerPDF(ruta: str):
  """
  Reads a PDF file and creates a SQL insert statement from the data it contains

  :param path: Path to the PDF file
  :return: None
  """
  
  with pdfplumber.open(ruta) as pdf:
    preTabla = pdf.pages[0].extract_table()
    tabla = []
    for fila in preTabla:
      if None in fila:
        fila.insert(fila.index(None), '')
        fila.remove(None)
      tabla.append(fila)
    columns = tabla[0]
    preValores = [fila for fila in tabla[1:]]
    valores = []
    for fila in preValores:
      preFila = []
      for valor in fila:
        preFila.append(str(int(valor)) if valor.isdigit() else "\""+str(valor)+"\"")
      valores.append(preFila)
  return format_sql_insert(table="tabla", columns=columns, values=valores)